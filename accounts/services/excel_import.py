import openpyxl
from django.db import transaction
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string
from accounts.models import StudentProfile, CompanyProfile
from students.models import Branch, Skill
from core.services import NotificationService
from io import BytesIO
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.urls import reverse

User = get_user_model()

class BulkImportService:
    @staticmethod
    def process_student_excel(file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'duplicates': 0,
            'errors': []
        }
        
        # Check Headers
        header_row = [cell.value for cell in sheet[1]]
        required_headers = ['Name', 'Email', 'Password', 'Enrollment Number', 'Branch', 'CGPA', 'Phone Number', 'Skills']
        if not all(h in header_row for h in required_headers):
            results['errors'].append({'row': 1, 'reason': f"Missing required headers. Expected: {required_headers}"})
            return results
            
        header_map = {h: i for i, h in enumerate(header_row)}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
                
            results['total'] += 1
            name = str(row[header_map['Name']] or '').strip()
            email = str(row[header_map['Email']] or '').strip()
            password = str(row[header_map['Password']] or '').strip()
            enrollment = str(row[header_map['Enrollment Number']] or '').strip()
            branch_code = str(row[header_map['Branch']] or '').strip()
            cgpa_val = row[header_map['CGPA']]
            phone = str(row[header_map['Phone Number']] or '').strip()
            skills_str = str(row[header_map['Skills']] or '').strip()
            
            if not email or not name or not password:
                results['failed'] += 1
                results['errors'].append({'row': row_idx, 'reason': "Missing Name, Email, or Password"})
                continue
                
            if User.objects.filter(email=email).exists():
                results['failed'] += 1
                results['duplicates'] += 1
                results['errors'].append({'row': row_idx, 'reason': f"Duplicate Email: {email}"})
                continue
                
            try:
                cgpa = float(cgpa_val) if cgpa_val is not None else 0.0
            except ValueError:
                results['failed'] += 1
                results['errors'].append({'row': row_idx, 'reason': "Invalid CGPA format"})
                continue
                
            try:
                with transaction.atomic():
                    first_name = name.split()[0]
                    last_name = " ".join(name.split()[1:]) if len(name.split()) > 1 else ""
                    
                    temp_password = get_random_string(10, 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789!@#$%^&*()')
                    
                    user = User.objects.create_user(
                        email=email,
                        password=temp_password,
                        first_name=first_name,
                        last_name=last_name,
                        is_student=True,
                        is_active=False,
                        force_password_change=True
                    )
                    
                    branch = None
                    if branch_code:
                        branch, _ = Branch.objects.get_or_create(code=branch_code.upper(), defaults={'name': branch_code})
                        
                    profile = StudentProfile.objects.create(
                        user=user,
                        enrollment_number=enrollment,
                        branch=branch,
                        cgpa=cgpa,
                        phone_number=phone
                    )
                    
                    if skills_str:
                        skill_names = [s.strip() for s in skills_str.split(',') if s.strip()]
                        for s_name in skill_names:
                            skill_obj, _ = Skill.objects.get_or_create(name=s_name)
                            profile.skills.add(skill_obj)
                            
                    NotificationService.create_and_send(
                        user=user,
                        message=f"Welcome to CareerConnect! Your temporary password is {temp_password}. Please login and change it immediately.",
                        email_subject="Welcome to CareerConnect - Login Instructions",
                        email_template="emails/bulk_import_welcome.html",
                        context={
                            'user_email': email,
                            'temp_password': temp_password,
                            'login_url': 'https://campus-placement-system-1.onrender.com/login/',
                            'name': name
                        }
                    )
                    
                    token_generator = PasswordResetTokenGenerator()
                    token = token_generator.make_token(user)
                    uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
                    verify_url = "https://campus-placement-system-1.onrender.com" + reverse('verify_email', kwargs={'uidb64': uidb64, 'token': token})
                    
                    NotificationService.create_and_send(
                        user=user,
                        message="Please verify your email address.",
                        email_subject="Verify Your CareerConnect Account",
                        email_template="emails/verify_email.html",
                        context={
                            'user_name': getattr(user, 'first_name', '') or user.email,
                            'verify_url': verify_url
                        }
                    )
                            
                    results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append({'row': row_idx, 'reason': str(e)})
                
        return results

    @staticmethod
    def process_company_excel(file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'duplicates': 0,
            'errors': []
        }
        
        header_row = [cell.value for cell in sheet[1]]
        required_headers = ['Company Name', 'Email', 'Password', 'Website', 'Industry', 'Contact Person']
        if not all(h in header_row for h in required_headers):
            results['errors'].append({'row': 1, 'reason': f"Missing required headers. Expected: {required_headers}"})
            return results
            
        header_map = {h: i for i, h in enumerate(header_row)}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
                
            results['total'] += 1
            company_name = str(row[header_map['Company Name']] or '').strip()
            email = str(row[header_map['Email']] or '').strip()
            password = str(row[header_map['Password']] or '').strip()
            website = str(row[header_map['Website']] or '').strip()
            industry = str(row[header_map['Industry']] or '').strip()
            contact_person = str(row[header_map['Contact Person']] or '').strip()
            
            if not email or not company_name or not password:
                results['failed'] += 1
                results['errors'].append({'row': row_idx, 'reason': "Missing Company Name, Email, or Password"})
                continue
                
            if User.objects.filter(email=email).exists():
                results['failed'] += 1
                results['duplicates'] += 1
                results['errors'].append({'row': row_idx, 'reason': f"Duplicate Email: {email}"})
                continue
                
            try:
                with transaction.atomic():
                    first_name = contact_person.split()[0] if contact_person else company_name
                    last_name = " ".join(contact_person.split()[1:]) if contact_person and len(contact_person.split()) > 1 else ""
                    
                    temp_password = get_random_string(10, 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789!@#$%^&*()')
                    
                    user = User.objects.create_user(
                        email=email,
                        password=temp_password,
                        first_name=first_name,
                        last_name=last_name,
                        is_company=True,
                        is_active=False,
                        force_password_change=True
                    )
                    
                    CompanyProfile.objects.create(
                        user=user,
                        company_name=company_name,
                        website=website,
                        industry=industry
                    )
                    
                    NotificationService.create_and_send(
                        user=user,
                        message=f"Welcome to CareerConnect! Your temporary password is {temp_password}. Please login and change it immediately.",
                        email_subject="Welcome to CareerConnect - Login Instructions",
                        email_template="emails/bulk_import_welcome.html",
                        context={
                            'user_email': email,
                            'temp_password': temp_password,
                            'login_url': 'https://campus-placement-system-1.onrender.com/login/',
                            'name': company_name
                        }
                    )
                    
                    token_generator = PasswordResetTokenGenerator()
                    token = token_generator.make_token(user)
                    uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
                    verify_url = "https://campus-placement-system-1.onrender.com" + reverse('verify_email', kwargs={'uidb64': uidb64, 'token': token})
                    
                    NotificationService.create_and_send(
                        user=user,
                        message="Please verify your email address.",
                        email_subject="Verify Your CareerConnect Account",
                        email_template="emails/verify_email.html",
                        context={
                            'user_name': getattr(user, 'first_name', '') or user.email,
                            'verify_url': verify_url
                        }
                    )
                    
                    results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append({'row': row_idx, 'reason': str(e)})
                
        return results

    @staticmethod
    def generate_sample_student_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        headers = ['Name', 'Email', 'Password', 'Enrollment Number', 'Branch', 'CGPA', 'Phone Number', 'Skills']
        ws.append(headers)
        ws.append(['John Doe', 'john@example.com', 'SecurePass123!', 'ENR123456', 'CSE', 8.5, '9876543210', 'Python, Django, React'])
        ws.append(['Jane Smith', 'jane@example.com', 'SecurePass123!', 'ENR654321', 'ECE', 9.1, '9876543211', 'C++, VLSI'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_sample_company_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies"
        headers = ['Company Name', 'Email', 'Password', 'Website', 'Industry', 'Contact Person']
        ws.append(headers)
        ws.append(['Google', 'hr@google.com', 'SecurePass123!', 'https://google.com', 'Technology', 'Sundar P.'])
        ws.append(['Acme Corp', 'careers@acmecorp.com', 'SecurePass123!', 'https://acmecorp.com', 'Manufacturing', 'Wile E.'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
